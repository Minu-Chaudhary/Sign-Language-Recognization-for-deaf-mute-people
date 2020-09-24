################################# library definition ######################################
import cv2 
import numpy as np
import math
import openpyxl

################################# defining func ######################################
def frame_read():
     ret, frame = cap.read()
     return frame

def crop_frame(frame):
    cv2.rectangle(frame, (100, 100), (300, 300), (0, 255, 0), 0)
    crop_image = frame[100:300, 100:300]
    return crop_image

def frame_pre_process(crop_image):
    blur = cv2.GaussianBlur(crop_image, (3, 3), 0)
    hsv = cv2.cvtColor(blur, cv2.COLOR_BGR2HSV)
    mask2 = cv2.inRange(hsv, np.array([2, 0, 0]), np.array([20, 255, 255]))
    kernel = np.ones((5, 5))
    dilation = cv2.dilate(mask2, kernel, iterations=1) 
    erosion = cv2.erode(dilation, kernel, iterations=1)
    filtered = cv2.GaussianBlur(erosion, (3, 3), 0)
    ret, thresh = cv2.threshold(filtered, 127, 255, 0)
    #for displaying threshold image
    cv2.imshow("threshold",thresh)
    cv2.imshow("filter",filtered)
    return thresh

def data_abstract(thresh, crop_image):
    contours, hierarchy = cv2.findContours(thresh.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    contour = max(contours, key=lambda x: cv2.contourArea(x))
    hull = cv2.convexHull(contour)
    areahull = cv2.contourArea(hull)
    areacnt = cv2.contourArea(contour)
    arearatio=((areahull-areacnt)/areacnt)*100
    
    drawing = np.zeros(crop_image.shape, np.uint8) 
    cv2.drawContours(drawing, [contour],-1,(0,255,0),0)
    cv2.drawContours(drawing, [hull],-1,(0,255,0),0)
    hull = cv2.convexHull(contour, returnPoints=False)
    defects = cv2.convexityDefects(contour,hull)
    return defects,contour,areacnt, arearatio, drawing

def analysis(defects, contour, crop_image, areacnt, arearatio):
    count_defects = 0
    for i in range(defects.shape[0]):
        s,e,f,d = defects[i,0]          #s:start, e:end, f:far, d:distant
        start = tuple(contour[s][0])
        end = tuple(contour[e][0])
        far = tuple(contour[f][0])
           
        a = math.sqrt((end[0] - start[0]) ** 2 + (end[1] - start[1]) ** 2)
        b = math.sqrt((far[0] - start[0]) ** 2 + (far[1] - start[1]) ** 2)
        c = math.sqrt((end[0] - far[0]) ** 2 + (end[1] - far[1]) ** 2)
        s = (a+b+c)/2
        ar = math.sqrt(s*(s-a)*(s-b)*(s-c))
        d = (2*ar)/a
        angle = (math.acos((b ** 2 + c ** 2 - a ** 2) / (2 * b * c)) * 180) / 3.14

        if angle <=90  and d >30 :
            count_defects +=1
            cv2.circle(crop_image, far, 3, [0,0,255], -1) # -1 so that boundary will be inside 
            cv2.line(crop_image, start, end, [0,255,0], 2)
            
    return count_defects,arearatio

def include_new_sign(count_defects,arearatio,sr_no):
     gesture_name = input("write gesture name:")
     sheet_obj.append([sr_no+1,gesture_name,count_defects,arearatio])
                
def  display_out(frame, drawing, crop_image, sign_name = "addition mode"):
    cv2.putText(frame, sign_name, (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1,(0,0,255),2)
    cv2.imshow("gesture", frame)
    all_image = np.hstack((drawing, crop_image))  
    cv2.imshow('contours', all_image)
 
################################# main func ######################################
if __name__ == "__main__":

    cap = cv2.VideoCapture(0)
    wb_obj = openpyxl.load_workbook("gesture2.xlsx")
    sheet_obj = wb_obj.active
    sr_no_max = sheet_obj.max_row     
    
    while(1):
        ip=input("Enter your choice: For Detection, enter D; For Addition, enter A:")
        if(ip == "A"):
             
             for i in range(5):
                  print("press a key 'n' after showing ur gesture")
                  while cap.isOpened():
                       try:
                            frame=frame_read()
                            crop_image=crop_frame(frame)
                            thresh=frame_pre_process(crop_image)
                            defects,contour,areacnt,arearatio,drawing = data_abstract(thresh, crop_image)
                            count_defects,arearatio = analysis(defects, contour, crop_image, areacnt, arearatio)
                            display_out(frame, drawing, crop_image)  
                       except:
                            pass
                              
                       if cv2.waitKey(1) == ord("n"): 
                            break
                  count_defects_t=0.0
                  arearatio_t=0.0
                  try:
                      defects,contour,areacnt,arearatio,drawing = data_abstract(thresh, crop_image)
                      count_defects,arearatio = analysis(defects, contour, crop_image, areacnt, arearatio)
                      count_defects_t = count_defects_t + count_defects
                      arearatio_t = arearatio_t + arearatio                     
                  except:
                      pass
             include_new_sign(count_defects,arearatio,sr_no_max)
             sr_no_max = sr_no_max+1
             wb_obj.save("gesture2.xlsx") 

        elif(ip == "D"):
             while cap.isOpened():
                  try:
                       frame=frame_read()
                       crop_image=crop_frame(frame)
                       thresh=frame_pre_process(crop_image)
                       defects,contour,areacnt,arearatio,drawing = data_abstract(thresh, crop_image)
                       count_defects,arearatio = analysis(defects, contour, crop_image, areacnt, arearatio)
                       
                  except:
                       pass
         
                    ################################################# comparing & detecting #######################################################
                  temp=1
                  temp_diff = 9999
                  for i in range(2,sr_no_max+1):
                      cell_obj = sheet_obj.cell(row =i,column =3)
                      cell_obj_ar = sheet_obj.cell(row =i,column =4)
                      a = cell_obj.value
                      b = cell_obj_ar.value

                      if((a == count_defects) and  (arearatio >= (b - b*0.2)) and  (arearatio <= (b + b*0.2))):
                          if(abs(b - arearatio) <= temp_diff ):
                              temp = i
                              temp_diff = abs(b - arearatio)
                              
                  if(temp == 1):
                       sign_name = "unknown sign"
                  else:
                       
                       cell_obj_n = sheet_obj.cell(row=temp,column = 2)
                       sign_name = cell_obj_n.value
                  display_out(frame, drawing, crop_image,sign_name)                
                  if cv2.waitKey(1) == ord("n"): 
                       break

        else:
             break
    cap.release()
    cv2.destroyAllWindows()
